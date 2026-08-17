from __future__ import annotations

import csv
import io
import re
from typing import Any

from openpyxl import load_workbook
import pandas as pd

ASIN_PATTERN = re.compile(r"(?<![A-Z0-9])([A-Z0-9]{10})(?![A-Z0-9])", re.I)


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def extract_asin(value: str) -> str:
    match = ASIN_PATTERN.search(_clean(value))
    return match.group(1).upper() if match else ""


def normalize_input_row(row: Any) -> dict[str, str]:
    if isinstance(row, str):
        value = row.strip()
        if "amazon." in value.lower() or value.startswith("http"):
            return {"asin": extract_asin(value), "url": value, "name": ""}
        asin = extract_asin(value)
        return {
            "asin": asin if len(value) == 10 else "",
            "url": "",
            "name": "" if asin == value.upper() else value,
        }
    if not isinstance(row, dict):
        return {"asin": "", "url": "", "name": ""}

    lowered = {str(key).strip().lower(): _clean(value) for key, value in row.items()}

    def first(*names: str) -> str:
        for name in names:
            if lowered.get(name):
                return lowered[name]
        return ""

    url = first("url", "product url", "amazon url", "web page url", "link")
    asin = first("asin", "amazon asin", "amazon id") or extract_asin(url)
    brand = first("brand", "manufacturer")
    model = first("model", "model number", "mpn", "manufacturer part number", "part number")
    upc = first("upc", "ean", "gtin", "gtin12", "gtin13", "barcode")
    name = first("name", "product name", "title", "product title", "item name", "description")
    if not name:
        name = " ".join(value for value in [brand, model, upc] if value)

    return {
        "asin": extract_asin(asin) or asin.upper(),
        "url": url,
        "name": name,
        "sku": first("sku", "item sku", "seller sku"),
        "upc": upc,
        "model": model,
        "brand": brand,
        "cost": first("cost", "unit cost", "purchase price", "price paid", "item cost"),
        "quantity": first("quantity", "qty", "inventory", "stock", "on hand"),
        "shipping_cost": first("shipping cost", "inbound shipping", "shipping"),
        "fees": first("fees", "fixed fees", "marketplace fees"),
        "fee_rate": first("fee rate", "fee %", "marketplace fee rate"),
        "category": first("category", "product category", "department"),
        "subcategory": first("subcategory", "sub category", "product type"),
        "condition": first("condition", "item condition"),
        "pack_count": first("pack count", "pack", "count", "units per pack"),
    }


def parse_upload(filename: str, data: bytes) -> tuple[list[dict[str, str]], list[str]]:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    if suffix == "csv":
        text = data.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = [{str(k): _clean(v) for k, v in row.items()} for row in reader]
        return rows, list(reader.fieldnames or [])

    if suffix in {"xlsx", "xlsm"}:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        matrix = list(sheet.iter_rows(values_only=True))
        if not matrix:
            return [], []
        headers = [
            _clean(value) or f"Column {index + 1}"
            for index, value in enumerate(matrix[0])
        ]
        rows = [
            dict(zip(headers, [_clean(value) for value in row]))
            for row in matrix[1:]
            if any(_clean(value) for value in row)
        ]
        return rows, headers

    if suffix == "xls":
        try:
            frame = pd.read_excel(io.BytesIO(data), dtype=str).fillna("")
            return frame.to_dict(orient="records"), [str(column) for column in frame.columns]
        except Exception as exc:
            raise ValueError(f"Could not read the XLS file: {exc}")

    raise ValueError("Unsupported file type. Upload CSV, XLSX, or XLS.")
